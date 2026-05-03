#!/usr/bin/env python3
"""diary JSON を visitors.xlsx / goods.xlsx に追記する"""

import argparse
import datetime
import json
from pathlib import Path

import openpyxl

# シート名
SHEET_VISITORS = "R7_入館者"
SHEET_GOODS = "R7_物販"

# R7_入館者 列番号（1-indexed）
V_DATE = 1        # 日付
V_CODE = 2        # 入館者コード
V_CATEGORY = 3    # 入館者分類（VLOOKUPキー）
V_DISCOUNT = 4    # 割引リスト
V_COUNT = 5       # 入場者数
V_UNIT_PRICE = 6  # 入館料単価（VLOOKUP数式）
V_TOTAL = 7       # 入館料合計（乗算数式）
V_EXHIBITION = 8  # 展覧会名

# R7_物販 列番号（1-indexed）
G_DATE = 1        # 日付
G_CODE = 2        # 図版コード
G_NAME = 3        # 図録・絵葉書（VLOOKUPキー）
G_COUNT = 4       # 販売数
G_UNIT_PRICE = 5  # 物販単価（VLOOKUP数式）
G_TOTAL = 6       # 物販合計（乗算数式）
G_EXHIBITION = 7  # 展覧会名

# 単価VLOOKUPの数式テンプレート
# リストシートはB3:C18（入場者分類→金額）、C列がVLOOKUPキー
_V_PRICE_FMT = '=IF(C{r}="","",IFERROR(VLOOKUP(C{r},リスト!$B$3:$C$18,2,FALSE),""))'
_V_TOTAL_FMT = '=IF(OR(E{r}="",F{r}=""),"",E{r}*F{r})'

# 図録リストシートはB3:C10（図録名→金額）、C列がVLOOKUPキー
_G_PRICE_FMT = '=IF(C{r}="","",IFERROR(VLOOKUP(C{r},図録リスト!$B$3:$C$10,2,FALSE),""))'
_G_TOTAL_FMT = '=IF(OR(D{r}="",E{r}=""),"",D{r}*E{r})'


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="diary JSON を Excel に追記")
    parser.add_argument("--json", required=True, help="diary_YYYY_MM.json のパス")
    parser.add_argument("--visitors", default="visitors.xlsx", help="visitors.xlsx のパス")
    parser.add_argument("--goods", default="goods.xlsx", help="goods.xlsx のパス")
    return parser.parse_args()


def find_next_empty_row(ws, date_col: int = 1) -> int:
    """A列がNullの最初の行番号を返す（データ末尾の次行）"""
    for row in ws.iter_rows(min_row=2):
        if row[date_col - 1].value is None:
            return row[0].row
    return ws.max_row + 1


def get_existing_dates(ws, date_col: int = 1) -> set[str]:
    """シートに既に存在する日付文字列（YYYY-MM-DD）の集合を返す"""
    dates: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[date_col - 1]
        if val is None:
            break
        if isinstance(val, (datetime.date, datetime.datetime)):
            dates.add(val.strftime("%Y-%m-%d"))
        else:
            dates.add(str(val)[:10])
    return dates


def _set_formula_if_empty(ws, row: int, col: int, formula: str) -> None:
    """セルが空の場合のみ数式を書き込む（既存数式を壊さない）"""
    if ws.cell(row=row, column=col).value is None:
        ws.cell(row=row, column=col).value = formula


def append_visitors(ws, day: dict, start_row: int) -> int:
    """1日分の visitors 配列を R7_入館者シートに追記する。追記行数を返す。"""
    date_val = datetime.date.fromisoformat(day["date"])
    exhibition = day.get("exhibition", "")
    row = start_row
    appended = 0

    for v in day.get("visitors", []):
        ws.cell(row=row, column=V_DATE).value = date_val
        ws.cell(row=row, column=V_CODE).value = v.get("code", "")
        ws.cell(row=row, column=V_CATEGORY).value = v.get("category", "")
        ws.cell(row=row, column=V_DISCOUNT).value = v.get("discount") or None
        ws.cell(row=row, column=V_COUNT).value = v.get("count")
        ws.cell(row=row, column=V_EXHIBITION).value = exhibition

        _set_formula_if_empty(ws, row, V_UNIT_PRICE, _V_PRICE_FMT.format(r=row))
        _set_formula_if_empty(ws, row, V_TOTAL, _V_TOTAL_FMT.format(r=row))

        row += 1
        appended += 1

    return appended


def append_goods(ws, day: dict, start_row: int) -> int:
    """1日分の sales 配列を R7_物販シートに追記する。追記行数を返す。"""
    date_val = datetime.date.fromisoformat(day["date"])
    exhibition = day.get("exhibition", "")
    row = start_row
    appended = 0

    for s in day.get("sales", []):
        ws.cell(row=row, column=G_DATE).value = date_val
        ws.cell(row=row, column=G_CODE).value = s.get("code", "")
        ws.cell(row=row, column=G_NAME).value = s.get("name", "")
        ws.cell(row=row, column=G_COUNT).value = s.get("count")
        ws.cell(row=row, column=G_EXHIBITION).value = exhibition

        _set_formula_if_empty(ws, row, G_UNIT_PRICE, _G_PRICE_FMT.format(r=row))
        _set_formula_if_empty(ws, row, G_TOTAL, _G_TOTAL_FMT.format(r=row))

        row += 1
        appended += 1

    return appended


def process_visitors(data: list[dict], excel_path: str) -> None:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[SHEET_VISITORS]
    existing = get_existing_dates(ws)
    next_row = find_next_empty_row(ws)
    total_rows = 0

    for day in data:
        date_str = day.get("date", "")
        if not date_str:
            continue
        if not day.get("visitors"):
            continue
        if date_str in existing:
            print(f"[SKIP] {date_str}: 入館者シート重複")
            continue
        n = append_visitors(ws, day, next_row)
        next_row += n
        total_rows += n
        print(f"[OK]   {date_str}: 入館者 {n}行追記")

    wb.save(excel_path)
    print(f"[SAVED] {excel_path}  ({total_rows}行追記)")


def process_goods(data: list[dict], excel_path: str) -> None:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[SHEET_GOODS]
    existing = get_existing_dates(ws)
    next_row = find_next_empty_row(ws)
    total_rows = 0

    for day in data:
        date_str = day.get("date", "")
        if not date_str:
            continue
        if not day.get("sales"):
            continue
        if date_str in existing:
            print(f"[SKIP] {date_str}: 物販シート重複")
            continue
        n = append_goods(ws, day, next_row)
        next_row += n
        total_rows += n
        print(f"[OK]   {date_str}: 物販 {n}行追記")

    wb.save(excel_path)
    print(f"[SAVED] {excel_path}  ({total_rows}行追記)")


def run(
    json_path: str,
    visitors_path: str = "visitors.xlsx",
    goods_path: str = "goods.xlsx",
) -> None:
    """GUIやスクリプトから直接呼び出せるエントリポイント。"""
    data: list[dict] = json.loads(Path(json_path).read_text(encoding="utf-8"))
    print(f"[INFO] JSON読込: {len(data)}日分  →  {json_path}")
    process_visitors(data, visitors_path)
    process_goods(data, goods_path)


def main() -> None:
    args = parse_args()
    run(args.json, args.visitors, args.goods)


if __name__ == "__main__":
    main()
